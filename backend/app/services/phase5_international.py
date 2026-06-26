from __future__ import annotations

import json
import subprocess
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from urllib.error import URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import (
    Competition,
    InternationalBacktestPrediction,
    InternationalBacktestRun,
    InternationalMatch,
    InternationalOdds,
    WorldCupMatch,
)
from app.services.backtest_engine import (
    TeamState,
    _brier,
    _dixon_coles_score_matrix,
    _draw_calibration_adjustment,
    _elo_probabilities,
    _form_probabilities,
    _normalize,
    _pick,
    _poisson_probabilities,
    _result,
    _update_elo,
    _update_team_state,
)


SOURCE_URL = "https://www.football-data.co.uk/WorldCup2026.xlsx"
COMPETITION_CODE = "WCQ_2026"
MODEL_VERSION = "international_value_v1"
RESULTS = ("Home Win", "Draw", "Away Win")
REGIONS = ("UEFA", "CONMEBOL")
UEFA_TEAMS = {
    "Albania", "Andorra", "Armenia", "Austria", "Azerbaijan", "Belarus", "Belgium",
    "Bosnia & Herzegovina", "Bulgaria", "Croatia", "Cyprus", "Czech Republic",
    "Denmark", "England", "Estonia", "Faroe Islands", "Finland", "France", "Georgia",
    "Germany", "Gibraltar", "Greece", "Hungary", "Iceland", "Ireland", "Israel",
    "Italy", "Kazakhstan", "Kosovo", "Latvia", "Liechtenstein", "Lithuania",
    "Luxembourg", "Malta", "Moldova", "Montenegro", "Netherlands", "North Macedonia",
    "Northern Ireland", "Norway", "Poland", "Portugal", "Romania", "San Marino",
    "Scotland", "Serbia", "Slovakia", "Slovenia", "Spain", "Sweden", "Switzerland",
    "Turkey", "Ukraine", "Wales",
}
CONMEBOL_TEAMS = {
    "Argentina", "Bolivia", "Brazil", "Chile", "Colombia", "Ecuador", "Paraguay",
    "Peru", "Uruguay", "Venezuela",
}
MODEL_WEIGHTS = {
    "elo": 0.35,
    "dixon_coles": 0.35,
    "form": 0.20,
    "market": 0.10,
}
EDGE_BUCKETS = (
    ("5%-10%", 0.05, 0.10),
    ("10%-15%", 0.10, 0.15),
    ("15%-20%", 0.15, 0.20),
    ("20%-25%", 0.20, 0.25),
    ("25%+", 0.25, None),
)


@dataclass
class Phase5Result:
    imported: dict[str, int]
    reports: dict[str, Path]
    metrics: dict[str, dict]


def _download_workbook(data_dir: Path) -> Path:
    data_dir.mkdir(parents=True, exist_ok=True)
    path = data_dir / "WorldCup2026.xlsx"
    request = Request(SOURCE_URL, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=60) as response:
            path.write_bytes(response.read())
    except (OSError, URLError):
        subprocess.run(
            ["curl.exe", "-L", "--ssl-no-revoke", "-o", str(path), SOURCE_URL],
            check=True,
            capture_output=True,
            text=True,
        )
    return path


def _region_for_match(home: str, away: str) -> str | None:
    if home in UEFA_TEAMS and away in UEFA_TEAMS:
        return "UEFA"
    if home in CONMEBOL_TEAMS and away in CONMEBOL_TEAMS:
        return "CONMEBOL"
    return None


def _competition(db: Session) -> Competition:
    competition = db.scalar(select(Competition).where(Competition.code == COMPETITION_CODE))
    if competition:
        return competition
    competition = Competition(
        code=COMPETITION_CODE,
        name="World Cup 2026 Qualifiers",
        region="International",
        source=SOURCE_URL,
    )
    db.add(competition)
    db.flush()
    return competition


def _market_probabilities(odds: InternationalOdds) -> dict[str, float]:
    implied = {
        "Home Win": 1 / odds.home_win_odds,
        "Draw": 1 / odds.draw_odds,
        "Away Win": 1 / odds.away_win_odds,
    }
    return _normalize(implied)


def _odds_for_pick(odds: InternationalOdds, pick: str) -> float:
    return {
        "Home Win": odds.home_win_odds,
        "Draw": odds.draw_odds,
        "Away Win": odds.away_win_odds,
    }[pick]


def _profit(pick: str, actual: str, odds: float, stake: float = 1.0) -> float:
    return stake * (odds - 1) if pick == actual else -stake


def _kelly_stake(probability: float, odds: float, fraction: float = 0.25) -> float:
    edge = probability * odds - 1
    raw_fraction = max(0.0, edge / max(odds - 1, 1e-9))
    return min(raw_fraction * fraction, 1.0)


def _max_drawdown(curve: list[dict]) -> float:
    peak = 0.0
    drawdown = 0.0
    for point in curve:
        value = point["value"]
        peak = max(peak, value)
        drawdown = min(drawdown, value - peak)
    return round(drawdown, 4)


def _roi_stats(rows: list[InternationalBacktestPrediction], kelly: bool = False) -> dict:
    stake_total = 0.0
    profit_total = 0.0
    hits = 0
    curve = []
    for row in rows:
        probability = {
            "Home Win": row.home_win_probability,
            "Draw": row.draw_probability,
            "Away Win": row.away_win_probability,
        }[row.best_value_pick]
        stake = _kelly_stake(probability, row.odds) if kelly else 1.0
        profit = _profit(row.best_value_pick, row.actual_result, row.odds, stake)
        stake_total += stake
        profit_total += profit
        hits += row.hit
        curve.append(
            {
                "match_id": row.match_id,
                "stake": round(stake, 4),
                "profit": round(profit, 4),
                "value": round(profit_total, 4),
            }
        )
    return {
        "matches": len(rows),
        "hits": hits,
        "hit_rate": round(hits / len(rows) * 100, 2) if rows else 0.0,
        "stake": round(stake_total, 4),
        "profit": round(profit_total, 4),
        "roi": round(profit_total / stake_total * 100, 2) if stake_total else None,
        "max_drawdown": _max_drawdown(curve),
        "equity_curve": curve,
    }


def import_phase5_matches(db: Session, workbook_path: Path) -> dict[str, int]:
    competition = _competition(db)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["WorldCup2026Qualifiers"]
    rows = list(sheet.iter_rows(values_only=True))
    header = {name: index for index, name in enumerate(rows[0])}
    imported = {"UEFA": 0, "CONMEBOL": 0}

    for row in rows[1:]:
        home = row[header["Home"]]
        away = row[header["Away"]]
        region = _region_for_match(home, away)
        if region not in REGIONS:
            continue
        home_score = row[header["HG"]]
        away_score = row[header["AG"]]
        home_avg = row[header["H_Avg"]]
        draw_avg = row[header["D_Avg"]]
        away_avg = row[header["A_Avg"]]
        if None in (home_score, away_score, home_avg, draw_avg, away_avg):
            continue
        match_date = row[header["Date"]]
        existing = db.scalar(
            select(InternationalMatch).where(
                InternationalMatch.competition_id == competition.id,
                InternationalMatch.match_date == match_date,
                InternationalMatch.home_team == home,
                InternationalMatch.away_team == away,
            )
        )
        match = existing or InternationalMatch(
            competition_id=competition.id,
            season="2026",
            qualifier_region=region,
            stage="World Cup Qualifier",
            match_date=match_date,
            home_team=home,
            away_team=away,
            home_score=int(home_score),
            away_score=int(away_score),
            result=_result(int(home_score), int(away_score)),
            source=SOURCE_URL,
        )
        match.home_score = int(home_score)
        match.away_score = int(away_score)
        match.result = _result(int(home_score), int(away_score))
        db.add(match)
        db.flush()

        odds = db.scalar(select(InternationalOdds).where(InternationalOdds.match_id == match.id))
        odds = odds or InternationalOdds(match_id=match.id)
        odds.source = SOURCE_URL
        odds.home_win_odds = float(home_avg)
        odds.draw_odds = float(draw_avg)
        odds.away_win_odds = float(away_avg)
        odds.home_win_max_odds = float(row[header["H_Max"]]) if row[header["H_Max"]] is not None else None
        odds.draw_max_odds = float(row[header["D_Max"]]) if row[header["D_Max"]] is not None else None
        odds.away_win_max_odds = float(row[header["A_Max"]]) if row[header["A_Max"]] is not None else None
        db.add(odds)
        if not existing:
            imported[region] += 1

    db.commit()
    return imported


def _seed_states_from_world_cup(db: Session, before_match: InternationalMatch, states: dict[str, TeamState]) -> None:
    rows = db.scalars(
        select(WorldCupMatch)
        .where(WorldCupMatch.match_date < before_match.match_date)
        .order_by(WorldCupMatch.match_date.asc(), WorldCupMatch.id.asc())
    ).all()
    for match in rows:
        home = states[match.home_team]
        away = states[match.away_team]
        _update_elo(home, away, match.home_score, match.away_score)
        _update_team_state(home, match.home_score, match.away_score, match.stage)
        _update_team_state(away, match.away_score, match.home_score, match.stage)


def _predict_match(match: InternationalMatch, odds: InternationalOdds, states: dict[str, TeamState]) -> tuple[dict, dict]:
    home = states[match.home_team]
    away = states[match.away_team]
    scores = _dixon_coles_score_matrix(home, away)
    market = _market_probabilities(odds)
    components = {
        "elo": _elo_probabilities(home, away),
        "dixon_coles": _poisson_probabilities(scores),
        "form": _form_probabilities(home, away),
        "market": market,
    }
    final = {result: 0.0 for result in RESULTS}
    for component, probabilities in components.items():
        for result in RESULTS:
            final[result] += probabilities[result] * MODEL_WEIGHTS[component]
    final = _normalize(final)
    final = _draw_calibration_adjustment(
        final,
        home,
        away,
        scores,
        match.stage,
        match.home_team,
        match.away_team,
        None,
    )
    return final, components


def _run_backtest_for_regions(db: Session, regions: tuple[str, ...], run_name: str) -> InternationalBacktestRun:
    competition = _competition(db)
    matches = db.execute(
        select(InternationalMatch, InternationalOdds)
        .join(InternationalOdds, InternationalOdds.match_id == InternationalMatch.id)
        .where(InternationalMatch.competition_id == competition.id)
        .where(InternationalMatch.qualifier_region.in_(regions))
        .order_by(InternationalMatch.match_date.asc(), InternationalMatch.id.asc())
    ).all()
    run = InternationalBacktestRun(
        run_name=run_name,
        model_version=MODEL_VERSION,
        competition_code=COMPETITION_CODE,
        region_filter=",".join(regions),
        sample_size=len(matches),
        metrics="{}",
    )
    db.add(run)
    db.flush()
    states: dict[str, TeamState] = defaultdict(TeamState)
    if matches:
        _seed_states_from_world_cup(db, matches[0][0], states)

    predictions: list[InternationalBacktestPrediction] = []
    for match, odds in matches:
        probabilities, components = _predict_match(match, odds, states)
        market = _market_probabilities(odds)
        edge_by_result = {result: probabilities[result] - market[result] for result in RESULTS}
        value_pick = max(RESULTS, key=lambda result: edge_by_result[result])
        value_edge = edge_by_result[value_pick]
        signal = "推荐" if value_edge >= 0.15 else "观察" if value_edge >= 0.05 else "跳过"
        price = _odds_for_pick(odds, value_pick)
        profit = _profit(value_pick, match.result, price)
        row = InternationalBacktestPrediction(
            run_id=run.id,
            match_id=match.id,
            predicted_result=_pick(probabilities),
            actual_result=match.result,
            home_win_probability=probabilities["Home Win"],
            draw_probability=probabilities["Draw"],
            away_win_probability=probabilities["Away Win"],
            market_home_probability=market["Home Win"],
            market_draw_probability=market["Draw"],
            market_away_probability=market["Away Win"],
            best_value_pick=value_pick,
            best_value_edge=value_edge,
            value_bet_signal=signal,
            odds=price,
            stake=1.0,
            profit=profit,
            hit=value_pick == match.result,
            component_probabilities=json.dumps(components, ensure_ascii=False),
        )
        db.add(row)
        predictions.append(row)
        home = states[match.home_team]
        away = states[match.away_team]
        _update_elo(home, away, match.home_score, match.away_score)
        _update_team_state(home, match.home_score, match.away_score, match.stage)
        _update_team_state(away, match.away_score, match.home_score, match.stage)

    metrics = _report_metrics(predictions, matches)
    run.metrics = json.dumps(metrics, ensure_ascii=False)
    db.commit()
    return run


def _bucket(rows: list[InternationalBacktestPrediction], lower: float, upper: float | None) -> list[InternationalBacktestPrediction]:
    return [row for row in rows if row.best_value_edge >= lower and (upper is None or row.best_value_edge < upper)]


def _report_metrics(predictions: list[InternationalBacktestPrediction], matches: list[tuple[InternationalMatch, InternationalOdds]]) -> dict:
    all_rows = list(predictions)
    edge_15 = [row for row in all_rows if row.best_value_edge >= 0.15]
    actual_by_match = {match.id: match.result for match, _ in matches}
    predicted_hits = sum(row.predicted_result == actual_by_match[row.match_id] for row in all_rows)
    brier = 0.0
    for row in all_rows:
        probabilities = {
            "Home Win": row.home_win_probability,
            "Draw": row.draw_probability,
            "Away Win": row.away_win_probability,
        }
        brier += _brier(probabilities, row.actual_result)
    buckets = {}
    for label, lower, upper in EDGE_BUCKETS:
        rows = _bucket(all_rows, lower, upper)
        buckets[label] = {
            "unit": _roi_stats(rows, kelly=False),
            "kelly": _roi_stats(rows, kelly=True),
        }
    return {
        "sample_size": len(all_rows),
        "prediction_hit_rate": round(predicted_hits / len(all_rows) * 100, 2) if all_rows else 0.0,
        "brier_score": round(brier / len(all_rows), 4) if all_rows else None,
        "edge_15": {
            "unit": _roi_stats(edge_15, kelly=False),
            "kelly": _roi_stats(edge_15, kelly=True),
        },
        "buckets": buckets,
    }


def _write_report(path: Path, title: str, metrics: dict, comparison: dict | None = None) -> None:
    edge = metrics["edge_15"]
    unit = edge["unit"]
    kelly = edge["kelly"]
    lines = [
        f"# {title}",
        "",
        "## 样本概览",
        "",
        f"- 样本量：{metrics['sample_size']}",
        f"- 胜平负预测命中率：{metrics['prediction_hit_rate']}%",
        f"- Brier Score：{metrics['brier_score']}",
        "",
        "## Edge >= 15%",
        "",
        f"- 场数：{unit['matches']}",
        f"- 命中率：{unit['hit_rate']}%",
        f"- 单位投注 ROI：{unit['roi']}%",
        f"- Kelly ROI：{kelly['roi']}%",
        f"- 最大回撤：{unit['max_drawdown']}",
        f"- 单位投注利润：{unit['profit']}",
        f"- Kelly 利润：{kelly['profit']}",
        "",
        "## Edge 分层",
        "",
        "| Edge区间 | 场数 | 命中率 | 单位投注ROI | Kelly ROI | 最大回撤 |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for label, payload in metrics["buckets"].items():
        unit_row = payload["unit"]
        kelly_row = payload["kelly"]
        lines.append(
            f"| {label} | {unit_row['matches']} | {unit_row['hit_rate']}% | "
            f"{unit_row['roi']}% | {kelly_row['roi']}% | {unit_row['max_drawdown']} |"
        )
    if comparison:
        lines.extend(
            [
                "",
                "## 与世界杯正赛样本对比",
                "",
                f"- 世界杯正赛 Edge>=15% 场数：{comparison['world_cup_edge_15_matches']}",
                f"- 世界杯正赛 Edge>=15% 单位投注 ROI：{comparison['world_cup_edge_15_roi']}%",
                f"- 当前样本 Edge>=15% 单位投注 ROI：{unit['roi']}%",
                f"- 判断：{comparison['verdict']}",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def _phase5_verdict(combined_metrics: dict) -> str:
    roi = combined_metrics["edge_15"]["unit"]["roi"]
    matches = combined_metrics["edge_15"]["unit"]["matches"]
    if matches < 30:
        return "Edge>=15% 样本不足，暂不能判断跨赛事稳定性"
    if roi is not None and roi > 0:
        return "Edge>=15% 在预选赛合并样本中仍为正 ROI，世界杯 ROI 存在跨赛事复现迹象"
    return "Edge>=15% 在预选赛合并样本中未能复现正 ROI，世界杯 ROI 暂不能跨赛事确认"


def run_phase5(db: Session, replace_previous: bool = True) -> Phase5Result:
    data_dir = Path(__file__).resolve().parents[2] / "data"
    reports_dir = Path(__file__).resolve().parents[2] / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = _download_workbook(data_dir)
    imported = import_phase5_matches(db, workbook_path)
    if replace_previous:
        existing_runs = db.scalars(
            select(InternationalBacktestRun).where(InternationalBacktestRun.model_version == MODEL_VERSION)
        ).all()
        for run in existing_runs:
            db.execute(delete(InternationalBacktestPrediction).where(InternationalBacktestPrediction.run_id == run.id))
            db.delete(run)
        db.commit()

    uefa_run = _run_backtest_for_regions(db, ("UEFA",), "Phase5 UEFA WCQ")
    conmebol_run = _run_backtest_for_regions(db, ("CONMEBOL",), "Phase5 CONMEBOL WCQ")
    combined_run = _run_backtest_for_regions(db, REGIONS, "Phase5 UEFA+CONMEBOL WCQ")
    metrics = {
        "UEFA": json.loads(uefa_run.metrics),
        "CONMEBOL": json.loads(conmebol_run.metrics),
        "COMBINED": json.loads(combined_run.metrics),
    }
    comparison = {
        "world_cup_edge_15_matches": 31,
        "world_cup_edge_15_roi": 39.61,
        "verdict": _phase5_verdict(metrics["COMBINED"]),
    }
    reports = {
        "UEFA": reports_dir / "phase5_uefa_report.md",
        "CONMEBOL": reports_dir / "phase5_conmebol_report.md",
        "COMBINED": reports_dir / "phase5_combined_report.md",
    }
    _write_report(reports["UEFA"], "Phase 5 UEFA 世界杯欧洲区预选赛报告", metrics["UEFA"], comparison)
    _write_report(reports["CONMEBOL"], "Phase 5 CONMEBOL 世界杯南美区预选赛报告", metrics["CONMEBOL"], comparison)
    _write_report(reports["COMBINED"], "Phase 5 合并样本报告", metrics["COMBINED"], comparison)
    return Phase5Result(imported=imported, reports=reports, metrics=metrics)
