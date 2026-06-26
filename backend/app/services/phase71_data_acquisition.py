from __future__ import annotations

import csv
import json
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import WorldCupMatch, WorldCupTeamProfile
from app.services.world_cup_specialist import (
    OFFICIAL_2026_QUALIFIED_TEAMS,
    PROFILE_FIELDS,
    calculate_team_upset_alert,
    calculate_world_cup_strength_score,
    import_team_profiles_csv,
    seed_2026_qualified_teams,
)


TRANSFERMARKT_PARTICIPANTS_URL = "https://www.transfermarkt.us/world-cup/teilnehmer/pokalwettbewerb/FIWC"
TRANSFERMARKT_RANKING_URL = "https://www.transfermarkt.us/statistik/weltrangliste?page={page}"
FOOTBALL_DATA_WORKBOOK = "https://www.football-data.co.uk/WorldCup2026.xlsx"
OUTPUT_CSV = "world_cup_team_profiles.csv"
OUTPUT_COVERAGE = "coverage_report.json"
SOURCE_TRANSFERMARKT = "Transfermarkt World Cup participants and FIFA world ranking pages"

NAME_ALIASES = {
    "USA": "United States",
    "United States": "USA",
    "IR Iran": "Iran",
    "Iran": "IR Iran",
    "Korea Republic": "South Korea",
    "South Korea": "Korea Republic",
    "Cote d'Ivoire": "Ivory Coast",
    "Côte d'Ivoire": "Ivory Coast",
    "Ivory Coast": "Cote d'Ivoire",
    "Curacao": "Curaçao",
    "Curaçao": "Curacao",
    "Turkiye": "Türkiye",
    "Türkiye": "Turkiye",
    "Bosnia and Herzegovina": "Bosnia-Herzegovina",
    "Bosnia-Herzegovina": "Bosnia and Herzegovina",
    "Cabo Verde": "Cape Verde",
    "Cape Verde": "Cabo Verde",
}


def _keys(name: str) -> set[str]:
    values = {name, NAME_ALIASES.get(name, name)}
    return {_normalize_name(value) for value in values}


def _normalize_name(name: str) -> str:
    text = (
        str(name)
        .replace(" ", " ")
        .replace("&", "and")
        .replace("U.S.", "USA")
        .strip()
    )
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    text = text.replace("'", "").replace("-", " ")
    return re.sub(r"\s+", " ", text).lower()


def _parse_money(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace("€", "").replace(",", "")
    if not text or text == "nan":
        return None
    multiplier = 1.0
    if text.endswith("bn"):
        multiplier = 1_000_000_000
        text = text[:-2]
    elif text.endswith("m"):
        multiplier = 1_000_000
        text = text[:-1]
    elif text.endswith("k"):
        multiplier = 1_000
        text = text[:-1]
    try:
        return round(float(text) * multiplier, 2)
    except ValueError:
        return None


def _safe_float(value: object) -> float | None:
    try:
        if value is None or str(value) == "nan":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_int(value: object) -> int | None:
    parsed = _safe_float(value)
    return int(parsed) if parsed is not None else None


def _read_transfermarkt_participants() -> dict[str, dict]:
    table = pd.read_html(TRANSFERMARKT_PARTICIPANTS_URL, storage_options={"User-Agent": "Mozilla/5.0"})[1]
    rows = {}
    for _, row in table.iterrows():
        country = str(row["Club"]).strip()
        rows[_normalize_name(country)] = {
            "squad_market_value_eur": _parse_money(row.get("Foreigners")),
            "average_age": _safe_float(row.get("Squad")),
            "world_cup_participations": _safe_int(row.get("&oslash-Age")),
        }
    return rows


def _read_transfermarkt_rankings(max_pages: int = 10) -> dict[str, dict]:
    rows = {}
    for page in range(1, max_pages + 1):
        try:
            table = pd.read_html(
                TRANSFERMARKT_RANKING_URL.format(page=page),
                storage_options={"User-Agent": "Mozilla/5.0"},
            )[1]
        except (IndexError, ValueError):
            continue
        for _, row in table.iterrows():
            country = str(row["Nation"]).strip()
            rows[_normalize_name(country)] = {
                "fifa_ranking": _safe_int(row.get("#")),
                "squad_market_value_eur": _parse_money(row.get("Total value")),
                "average_age": _safe_float(row.get("Avg. age")),
            }
    return rows


def _recent_two_year_ratings(data_dir: Path) -> dict[str, float]:
    workbook_path = data_dir / "WorldCup2026.xlsx"
    if not workbook_path.exists():
        return {}
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["WorldCup2026Qualifiers"]
    rows = list(sheet.iter_rows(values_only=True))
    header = {name: index for index, name in enumerate(rows[0])}
    stats: dict[str, dict[str, int]] = {}
    for row in rows[1:]:
        home = row[header["Home"]]
        away = row[header["Away"]]
        home_goals = row[header["HG"]]
        away_goals = row[header["AG"]]
        if None in (home, away, home_goals, away_goals):
            continue
        home_key = _normalize_name(str(home))
        away_key = _normalize_name(str(away))
        stats.setdefault(home_key, {"matches": 0, "points": 0})
        stats.setdefault(away_key, {"matches": 0, "points": 0})
        stats[home_key]["matches"] += 1
        stats[away_key]["matches"] += 1
        if home_goals > away_goals:
            stats[home_key]["points"] += 3
        elif home_goals < away_goals:
            stats[away_key]["points"] += 3
        else:
            stats[home_key]["points"] += 1
            stats[away_key]["points"] += 1
    return {
        team: round(value["points"] / (value["matches"] * 3), 4)
        for team, value in stats.items()
        if value["matches"]
    }


def _last_world_cup_finish(db: Session) -> dict[str, str]:
    rows = db.scalars(select(WorldCupMatch).where(WorldCupMatch.tournament_year == 2022)).all()
    stages: dict[str, set[str]] = {}
    for match in rows:
        stages.setdefault(_normalize_name(match.home_team), set()).add(match.stage)
        stages.setdefault(_normalize_name(match.away_team), set()).add(match.stage)
    priority = [
        ("Final", "Finalist"),
        ("Semi-finals", "Semi-final"),
        ("Semi", "Semi-final"),
        ("Quarter-finals", "Quarter-final"),
        ("Quarter", "Quarter-final"),
        ("Round of 16", "Round of 16"),
        ("Group", "Group stage"),
    ]
    finishes = {}
    for team, team_stages in stages.items():
        text = "Qualified 2022"
        for token, label in priority:
            if any(token in stage for stage in team_stages):
                text = label
                break
        finishes[team] = text
    return finishes


def _lookup(source: dict[str, dict] | dict[str, float] | dict[str, str], country: str):
    keys = _keys(country)
    for alias_key, alias_value in NAME_ALIASES.items():
        if _normalize_name(alias_key) in keys or _normalize_name(alias_value) in keys:
            keys.add(_normalize_name(alias_key))
            keys.add(_normalize_name(alias_value))
    for key in keys:
        if key in source:
            return source[key]
    return None


def _coverage(row: dict) -> tuple[int, list[str]]:
    required = [
        "fifa_ranking",
        "elo_rating",
        "squad_market_value_eur",
        "average_age",
        "total_caps",
        "average_caps",
        "coach",
        "last_world_cup_finish",
        "recent_two_year_rating",
    ]
    missing = [field for field in required if row.get(field) in (None, "")]
    return len(required) - len(missing), missing


def _write_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = [
        "tournament_year",
        "team_name",
        "country",
        "confederation",
        "fifa_ranking",
        "elo_rating",
        "squad_market_value_eur",
        "average_age",
        "total_caps",
        "average_caps",
        "world_cup_history_score",
        "recent_two_year_rating",
        "projected_starting_xi",
        "key_injuries",
        "coach",
        "last_world_cup_finish",
        "source",
        "captured_at",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def run_phase71_data_acquisition(db: Session) -> dict:
    data_dir = Path(__file__).resolve().parents[2] / "data"
    reports_dir = Path(__file__).resolve().parents[2] / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    data_dir.mkdir(parents=True, exist_ok=True)

    seed_2026_qualified_teams(db)
    participants = _read_transfermarkt_participants()
    rankings = _read_transfermarkt_rankings()
    recent = _recent_two_year_ratings(data_dir)
    finishes = _last_world_cup_finish(db)

    rows = []
    for country, confederation in OFFICIAL_2026_QUALIFIED_TEAMS:
        participant = _lookup(participants, country) or {}
        ranking = _lookup(rankings, country) or {}
        recent_rating = _lookup(recent, country)
        last_finish = _lookup(finishes, country) or "Did not qualify 2022"
        row = {
            "tournament_year": 2026,
            "team_name": country,
            "country": country,
            "confederation": confederation,
            "fifa_ranking": ranking.get("fifa_ranking"),
            "elo_rating": None,
            "squad_market_value_eur": ranking.get("squad_market_value_eur") or participant.get("squad_market_value_eur"),
            "average_age": ranking.get("average_age") or participant.get("average_age"),
            "total_caps": None,
            "average_caps": None,
            "world_cup_history_score": None,
            "recent_two_year_rating": recent_rating,
            "projected_starting_xi": "[]",
            "key_injuries": "[]",
            "coach": "",
            "last_world_cup_finish": last_finish,
            "source": SOURCE_TRANSFERMARKT + "; Football-Data WorldCup2026Qualifiers; local WorldCup 2022 results",
            "captured_at": "",
        }
        rows.append(row)

    csv_path = data_dir / OUTPUT_CSV
    coverage_path = reports_dir / OUTPUT_COVERAGE
    _write_csv(csv_path, rows)
    import_summary = import_team_profiles_csv(db, csv_path)

    profiles = db.scalars(
        select(WorldCupTeamProfile)
        .where(WorldCupTeamProfile.tournament_year == 2026)
        .order_by(WorldCupTeamProfile.country.asc())
    ).all()
    for profile in profiles:
        profile.world_cup_strength_score = calculate_world_cup_strength_score(profile)
        profile.upset_alert_score = calculate_team_upset_alert(profile)
        db.add(profile)
    db.commit()

    coverage_rows = []
    complete = 0
    field_total = 0
    field_present = 0
    for row in rows:
        present, missing = _coverage(row)
        field_total += 9
        field_present += present
        complete += int(not missing)
        coverage_rows.append(
            {
                "country": row["country"],
                "confederation": row["confederation"],
                "present_fields": present,
                "missing_fields": missing,
            }
        )
    coverage = {
        "total_teams": len(rows),
        "completed_teams": complete,
        "team_completion_rate": round(complete / len(rows) * 100, 2) if rows else 0.0,
        "field_coverage_rate": round(field_present / field_total * 100, 2) if field_total else 0.0,
        "target_field_coverage_rate": 90.0,
        "target_reached": field_present / field_total >= 0.90 if field_total else False,
        "missing_fields": coverage_rows,
        "source_status": {
            "transfermarkt_participants_rows": len(participants),
            "transfermarkt_ranking_rows": len(rankings),
            "recent_two_year_rating_rows": len(recent),
            "elo_rating": "missing: no stable machine-readable source confirmed",
            "coach": "missing: not collected from a stable source in this run",
            "caps": "missing: no stable machine-readable source confirmed",
        },
        "import_summary": {
            "imported": import_summary.imported,
            "updated": import_summary.updated,
            "skipped": import_summary.skipped,
        },
        "csv_path": str(csv_path),
    }
    coverage_path.write_text(json.dumps(coverage, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path = reports_dir / "phase71_data_coverage_report.md"
    lines = [
        "# Phase 7.1 World Cup Data Coverage Report",
        "",
        "## 范围",
        "",
        "- 停止新模型、新回测、新 ROI 模块。",
        "- 只采集 2026 世界杯 48 支球队真实数据。",
        "- 禁止 mock、随机、赛后泄露数据。",
        "",
        "## 输出文件",
        "",
        f"- CSV：{csv_path}",
        f"- Coverage JSON：{coverage_path}",
        "",
        "## 覆盖率",
        "",
        f"- 总球队数：{coverage['total_teams']}",
        f"- 已完成球队数：{coverage['completed_teams']}",
        f"- 球队完整覆盖率：{coverage['team_completion_rate']}%",
        f"- 字段覆盖率：{coverage['field_coverage_rate']}%",
        f"- 90% 目标是否达到：{coverage['target_reached']}",
        "",
        "## 数据源状态",
        "",
        f"- Transfermarkt 参赛队/身价/年龄/FIFA Ranking 行数：{len(rankings)}",
        f"- Football-Data 近两年评级可用球队数：{len(recent)}",
        "- ELO、coach、caps 本轮未找到稳定机器可读来源，保持缺失。",
        "",
        "## 缺失字段明细",
        "",
        "| 国家 | 大洲 | 已采字段数 | 缺失字段 |",
        "|---|---|---:|---|",
    ]
    for row in coverage_rows:
        lines.append(
            f"| {row['country']} | {row['confederation']} | {row['present_fields']} | {', '.join(row['missing_fields'])} |"
        )
    report_path.write_text("\n".join(lines), encoding="utf-8")
    coverage["coverage_path"] = str(coverage_path)
    coverage["report_path"] = str(report_path)
    return coverage
